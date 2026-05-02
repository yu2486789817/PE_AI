import openpyxl
from openpyxl.styles import Font, Alignment
import random

def generate_teacher_data(n):
    """
    生成n条教师信息
    """
    # 常见的中国姓氏
    surnames = [
        '王', '李', '张', '刘', '陈', '杨', '赵', '黄', '周', '吴',
        '徐', '孙', '胡', '朱', '高', '林', '何', '郭', '马', '罗',
        '梁', '宋', '郑', '谢', '韩', '唐', '冯', '于', '董', '萧',
        '程', '曹', '袁', '邓', '许', '傅', '沈', '曾', '彭', '吕',
        '苏', '卢', '蒋', '蔡', '贾', '丁', '魏', '薛', '叶', '阎'
    ]
    
    # 常见的中文名
    given_names = [
        '伟', '芳', '娜', '秀英', '敏', '静', '丽', '强', '磊', '军',
        '洋', '勇', '艳', '杰', '娟', '涛', '明', '超', '秀兰', '霞',
        '平', '刚', '桂英', '建华', '文', '华', '金凤', '素英', '建国', '德华',
        '秀珍', '志强', '秀荣', '丽娟', '建军', '春梅', '海燕', '雪梅', '美玲', '翠平',
        '小红', '小明', '小华', '小芳', '小军', '小丽', '小强', '小伟', '小杰', '小娟',
        '子涵', '浩然', '子轩', '欣怡', '佳怡', '俊杰', '子豪', '欣然', '佳豪', '佳欣'
    ]
    
    data = []
    for i in range(n):
        # 生成教师ID，格式为1980000-2025999
        student_id = f"{random.randint(1980000, 2025999)}"
        
        # 随机生成姓名
        surname = random.choice(surnames)
        given_name = random.choice(given_names)
        name = surname + given_name
        
        data.append([student_id, name])
    
    return data

def create_excel_file(data, filename='教师信息表.xlsx'):
    """
    创建Excel文件
    """
    # 创建工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '教师信息'
    
    # 设置表头
    headers = ['ID', 'name']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 填充数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    sheet.column_dimensions['A'].width = 15  # 学号列
    sheet.column_dimensions['B'].width = 15  # 姓名列
    
    # 保存文件
    workbook.save(filename)
    print(f"Excel文件 '{filename}' 已生成成功！")

def main():
    try:
        n = int(input("请输入要生成的教师人数: "))
        if n <= 0:
            print("请输入一个正整数！")
            return

        print(f"正在生成 {n} 条教师信息...")
        student_data = generate_teacher_data(n)
        create_excel_file(student_data)
        
    except ValueError:
        print("输入错误，请输入一个有效的数字！")
    except Exception as e:
        print(f"程序执行出错: {e}")

if __name__ == "__main__":
    main()